from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import main
from app.modules import excel


def _preparar_app(monkeypatch, tmp_path):
    padrones_dir = tmp_path / "padrones"
    salidas_dir = tmp_path / "salidas"
    uploads_dir = tmp_path / "uploads"
    padrones_dir.mkdir()
    salidas_dir.mkdir()
    uploads_dir.mkdir()
    monkeypatch.setattr(main, "PADRONES_DIR", padrones_dir)
    monkeypatch.setattr(main, "SALIDAS_DIR", salidas_dir)
    monkeypatch.setattr(main, "UPLOADS_DIR", uploads_dir)
    monkeypatch.delenv("BASIC_AUTH_USER", raising=False)
    monkeypatch.delenv("BASIC_AUTH_PASS", raising=False)
    monkeypatch.setattr(main.afip_arca, "consultar_constancia", lambda cuit: {
        "modo": "demo",
        "encontrado": True,
        "razon_social": "PROVEEDOR TEST SA",
        "tipo_persona": "JURIDICA",
        "estado_clave": "ACTIVO",
        "condicion_iva": "RESPONSABLE INSCRIPTO (ACTIVO)",
        "condicion_ganancias": "GANANCIAS SOCIEDADES (ACTIVO)",
        "inscripciones_iibb": {
            "regimen": "Convenio Multilateral",
            "jurisdicciones": ["CABA"],
            "impuestos": [{"descripcion": "CONVENIO MULTILATERAL", "estado": "ACTIVO"}],
        },
        "domicilio_fiscal": "Av. Test 123, CABA",
        "actividad_principal": "Servicios",
        "en_apoc": False,
    })
    monkeypatch.setattr(main.padrones, "consultar_todos", lambda cuit, path: {
        "CABA": {"status": "inscripto", "detalle": "Padrón vigente.", "nombre": "CABA / AGIP", "prioridad": "P1"},
        "Cordoba": {"status": "no_disponible", "detalle": "Falta archivo.", "nombre": "Córdoba", "prioridad": "P2"},
    })
    monkeypatch.setattr(main.fuentes_online, "consultar_todas", lambda cuit: {})
    monkeypatch.setattr(main.georef, "normalizar_provincia", lambda domicilio: {"provincia": "CABA"})
    return TestClient(main.app), salidas_dir


def test_api_validar_incluye_regimenes_aplicables(monkeypatch, tmp_path):
    client, _ = _preparar_app(monkeypatch, tmp_path)

    r = client.post("/api/validar", json={"cuits": ["33-69345023-9"]})

    assert r.status_code == 200
    resultado = r.json()["resultados"][0]
    mapa = resultado["regimenes_aplicables"]
    por_id = {item["id"]: item for item in mapa["items"]}
    assert mapa["resumen"]["total"] >= 1
    assert por_id["agip_caba_iibb_ret_per"]["clasificacion"] == "aplicable"
    assert "alicuota" not in por_id["agip_caba_iibb_ret_per"]
    assert por_id["agip_caba_iibb_ret_per"]["evidencia_requerida"]
    assert por_id["agip_caba_iibb_ret_per"]["proxima_accion"]


def test_excel_agrega_hoja_regimenes(tmp_path):
    salida = tmp_path / "validacion.xlsx"
    resultados = [{
        "cuit": "33-69345023-9",
        "valido": True,
        "tipo_persona": "JURIDICA",
        "mensaje_validador": "CUIT válido",
        "timestamp": "2026-05-26 10:00:00",
        "modo_afip": "demo",
        "afip": {"razon_social": "PROVEEDOR TEST SA", "estado_clave": "ACTIVO", "en_apoc": False},
        "padrones": {},
        "fuentes_online": {},
        "matriz_tributaria": {"items": [], "alertas": [], "resumen": "Sin matriz."},
        "georef": {"provincia": "CABA"},
        "decision_fiscal": {"estado": "APROBABLE", "label": "Aprobable", "motivos": [], "recomendaciones": []},
        "regimenes_aplicables": {
            "items": [{
                "id": "agip_caba_iibb_ret_per",
                "nombre": "AGIP/CABA · Padrón regímenes generales",
                "nivel": "provincial",
                "jurisdiccion": "CABA",
                "tipo": "retencion_percepcion",
                "clasificacion": "aplicable",
                "evidencia_requerida": "archivo_original_hash_manifest_preview_acuse",
                "proxima_accion": "Conservar evidencia.",
                "motivos": ["Padrón vigente."],
            }],
            "resumen": {"total": 1},
        },
    }]

    excel.generar(resultados, salida)

    wb = load_workbook(salida)
    assert "Regímenes" in wb.sheetnames
    ws = wb["Regímenes"]
    assert ws["A1"].value == "Mapa de obligaciones potenciales por proveedor"
    assert ws["C4"].value == "AGIP/CABA · Padrón regímenes generales"
    assert ws["G4"].value == "aplicable"

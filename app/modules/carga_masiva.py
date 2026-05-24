"""Lectura de CUITs desde archivos Excel/CSV para validación masiva."""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from openpyxl import load_workbook

CUIT_HEADERS = {"cuit", "cuil", "cuit_cuil", "cuit/cuil", "CUIT", "CUIL", "CUIT CUIL", "CUIT/CUIL"}


def _leer_texto(path: Path) -> str:
    data = path.read_bytes()
    for enc in ("utf-8-sig", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _norm_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _extraer_cuit(value) -> str | None:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return digits if len(digits) == 11 else None


def _detectar_columna(headers: list[str], preferida: str | None = None) -> int:
    if preferida:
        pref = _norm_header(preferida)
        for i, h in enumerate(headers):
            if _norm_header(h) == pref:
                return i
        raise ValueError(f"No encontré la columna indicada: {preferida}")
    for i, h in enumerate(headers):
        nh = _norm_header(h)
        if nh in {"cuit", "cuil", "cuit_cuil", "cuitcuil", "numero_cuit", "nro_cuit"}:
            return i
    raise ValueError("No pude detectar columna CUIT/CUIL. Indicá el nombre de columna manualmente.")


def _leer_csv(path: Path, columna: str | None) -> list[str]:
    texto = _leer_texto(path)
    try:
        dialect = csv.Sniffer().sniff(texto[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(texto), dialect=dialect)
    rows = list(reader)
    if not rows:
        return []
    col = _detectar_columna([str(x) for x in rows[0]], columna)
    cuits = []
    for row in rows[1:]:
        if col < len(row):
            cuit = _extraer_cuit(row[col])
            if cuit:
                cuits.append(cuit)
    return cuits


def _leer_xlsx(path: Path, columna: str | None, sheet: str | None) -> list[str]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows[:20]):
        vals = ["" if v is None else str(v).strip() for v in row]
        try:
            _detectar_columna(vals, columna)
            header_idx = i
            break
        except ValueError:
            continue
    if header_idx is None:
        raise ValueError("No pude detectar fila de cabecera con columna CUIT/CUIL.")
    headers = ["" if v is None else str(v).strip() for v in rows[header_idx]]
    col = _detectar_columna(headers, columna)
    cuits = []
    for row in rows[header_idx + 1:]:
        if col < len(row):
            cuit = _extraer_cuit(row[col])
            if cuit:
                cuits.append(cuit)
    return cuits


def leer_cuits(path: Path, columna: str | None = None, sheet: str | None = None) -> list[str]:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm"}:
        cuits = _leer_xlsx(path, columna, sheet)
    elif ext in {".csv", ".txt", ".tsv", ".psv"}:
        cuits = _leer_csv(path, columna)
    else:
        raise ValueError("Formato no soportado. Usá Excel o CSV/TXT.")
    dedup = []
    seen = set()
    for c in cuits:
        if c not in seen:
            seen.add(c)
            dedup.append(c)
    return dedup

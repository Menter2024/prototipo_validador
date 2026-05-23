"""Generador de Excel consolidado con los resultados de la validación."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path


HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, color="0F6FC6", size=16)
SUBTITLE_FONT = Font(name="Arial", italic=True, color="595959", size=10)
BASE_FONT = Font(name="Arial", size=10)

OK_FILL = PatternFill("solid", start_color="D1FADF")
WARN_FILL = PatternFill("solid", start_color="FEF3C7")
ERR_FILL = PatternFill("solid", start_color="FECDD3")
NEUT_FILL = PatternFill("solid", start_color="F1F5F9")

thin = Side(border_style="thin", color="D0D7DE")
ALL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _hdr(cell, text):
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = ALL_BORDER


def _cell(cell, text, status=None):
    cell.value = text
    cell.font = BASE_FONT
    cell.alignment = LEFT
    cell.border = ALL_BORDER
    if status == "ok":
        cell.fill = OK_FILL
    elif status == "warn":
        cell.fill = WARN_FILL
    elif status == "error":
        cell.fill = ERR_FILL


def generar(resultados: list, ruta_salida: Path) -> Path:
    """Genera un Excel multi-hoja a partir de la lista de resultados.

    Cada `resultado` es un dict con:
      - cuit, cuit_limpio, valido, validador (str), modo_afip
      - afip (dict con razon_social, etc.)
      - padrones (dict provincia -> {status, detalle})
      - georef (dict)
      - timestamp
    """
    wb = Workbook()

    # ---- Hoja 1: Resumen
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = "Validación Automatizada de Alta de Proveedores"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Generado por Menter S.A.S. · {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    total = len(resultados)
    total_validos = sum(1 for r in resultados if r.get("valido"))
    total_live = sum(1 for r in resultados if r.get("modo_afip") == "live")
    total_observados = 0
    for r in resultados:
        afip = r.get("afip", {})
        observado = (
            (not r.get("valido"))
            or afip.get("en_apoc")
            or (afip.get("estado_clave") not in (None, "—", "ACTIVO"))
            or any(p.get("status") == "inscripto" for p in r.get("padrones", {}).values())
        )
        total_observados += 1 if observado else 0

    kpis = [
        ("Procesados", total),
        ("CUIT válidos", total_validos),
        ("AFIP live", total_live),
        ("Observados", total_observados),
    ]
    for i, (label, value) in enumerate(kpis, start=1):
        cell = ws.cell(row=3, column=i)
        cell.value = f"{label}: {value}"
        cell.font = Font(name="Arial", bold=True, color="1F3864", size=10)
        cell.fill = NEUT_FILL
        cell.border = ALL_BORDER
        cell.alignment = CENTER

    headers = ["CUIT", "Razón Social", "Validez DV", "Estado AFIP", "Modo", "Hallazgos"]
    for i, h in enumerate(headers):
        _hdr(ws.cell(row=5, column=i + 1), h)

    for ri, r in enumerate(resultados):
        row = 6 + ri
        afip = r.get("afip", {})
        # Computamos hallazgos
        hallazgos = []
        if not r.get("valido", False):
            hallazgos.append("CUIT con DV inválido")
        if afip.get("en_apoc"):
            hallazgos.append("Figura en base APOC")
        if afip.get("estado_clave") not in (None, "—") and afip.get("estado_clave") != "ACTIVO":
            hallazgos.append(f"Estado AFIP: {afip.get('estado_clave')}")
        for prov, p in r.get("padrones", {}).items():
            if p.get("status") == "inscripto":
                hallazgos.append(f"Inscripto IIBB {prov}")

        status_resumen = "error" if any("APOC" in h or "DV" in h for h in hallazgos) else (
            "warn" if hallazgos else "ok"
        )
        _cell(ws.cell(row=row, column=1), r.get("cuit", "—"))
        _cell(ws.cell(row=row, column=2), afip.get("razon_social", "—"))
        _cell(ws.cell(row=row, column=3), "Válido" if r.get("valido") else "Inválido", "ok" if r.get("valido") else "error")
        _cell(ws.cell(row=row, column=4), afip.get("estado_clave", "—"))
        _cell(ws.cell(row=row, column=5), r.get("modo_afip", "demo").upper())
        _cell(ws.cell(row=row, column=6), " · ".join(hallazgos) if hallazgos else "Sin hallazgos", status_resumen)

    for col, w in zip("ABCDEF", [18, 36, 12, 18, 10, 50]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[5].height = 22
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:F{5 + max(len(resultados), 1)}"

    # ---- Hoja 2: AFIP / ARCA
    s2 = wb.create_sheet("AFIP-ARCA")
    s2["A1"] = "Datos de AFIP / ARCA"
    s2["A1"].font = TITLE_FONT
    s2.merge_cells("A1:I1")

    hdrs = ["CUIT", "Razón Social", "Tipo", "Estado", "Condición IVA",
            "Condición Ganancias", "Domicilio Fiscal", "Actividad", "En APOC"]
    for i, h in enumerate(hdrs):
        _hdr(s2.cell(row=3, column=i + 1), h)

    for ri, r in enumerate(resultados):
        afip = r.get("afip", {})
        row = 4 + ri
        en_apoc = afip.get("en_apoc")
        apoc_text = "SÍ" if en_apoc else ("NO" if en_apoc is False else "—")
        apoc_status = "error" if en_apoc else "ok" if en_apoc is False else None
        vals = [
            r.get("cuit", "—"),
            afip.get("razon_social", "—"),
            afip.get("tipo_persona", "—"),
            afip.get("estado_clave", "—"),
            afip.get("condicion_iva", "—"),
            afip.get("condicion_ganancias", "—"),
            afip.get("domicilio_fiscal", "—"),
            afip.get("actividad_principal", "—"),
            apoc_text,
        ]
        for i, v in enumerate(vals):
            st = apoc_status if i == 8 else None
            _cell(s2.cell(row=row, column=i + 1), v, st)

    for col, w in zip("ABCDEFGHI", [18, 32, 12, 12, 22, 18, 36, 30, 10]):
        s2.column_dimensions[col].width = w

    # ---- Hoja 3: Padrones IIBB
    s3 = wb.create_sheet("Padrones IIBB")
    s3["A1"] = "Resultado de la consulta a padrones provinciales"
    s3["A1"].font = TITLE_FONT

    # Cabeceras dinámicas: una columna por provincia
    if resultados:
        provincias = list(resultados[0].get("padrones", {}).keys())
    else:
        provincias = []
    s3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(provincias))

    hdrs3 = ["CUIT", "Razón Social"] + provincias
    for i, h in enumerate(hdrs3):
        _hdr(s3.cell(row=3, column=i + 1), h)

    for ri, r in enumerate(resultados):
        row = 4 + ri
        _cell(s3.cell(row=row, column=1), r.get("cuit", "—"))
        _cell(s3.cell(row=row, column=2), r.get("afip", {}).get("razon_social", "—"))
        for i, prov in enumerate(provincias):
            p = r.get("padrones", {}).get(prov, {})
            st = {"inscripto": "warn", "no_inscripto": "ok", "no_disponible": None}.get(p.get("status"))
            _cell(s3.cell(row=row, column=3 + i), p.get("detalle", "—"), st)

    s3.column_dimensions["A"].width = 18
    s3.column_dimensions["B"].width = 32
    for i in range(len(provincias)):
        s3.column_dimensions[get_column_letter(3 + i)].width = 38
    s3.row_dimensions[3].height = 30

    # ---- Hoja 4: Trazabilidad
    s4 = wb.create_sheet("Trazabilidad")
    s4["A1"] = "Trazabilidad de la consulta"
    s4["A1"].font = TITLE_FONT
    s4.merge_cells("A1:E1")

    hdrs4 = ["CUIT", "Timestamp", "Modo AFIP", "Mensaje validador", "Provincia normalizada"]
    for i, h in enumerate(hdrs4):
        _hdr(s4.cell(row=3, column=i + 1), h)

    for ri, r in enumerate(resultados):
        row = 4 + ri
        vals = [
            r.get("cuit", "—"),
            r.get("timestamp", "—"),
            r.get("modo_afip", "demo").upper(),
            r.get("mensaje_validador", "—"),
            r.get("georef", {}).get("provincia", "—") or "—",
        ]
        for i, v in enumerate(vals):
            _cell(s4.cell(row=row, column=i + 1), v)

    for col, w in zip("ABCDE", [18, 22, 12, 50, 28]):
        s4.column_dimensions[col].width = w

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)
    return ruta_salida

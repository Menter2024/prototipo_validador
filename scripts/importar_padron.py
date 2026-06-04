#!/usr/bin/env python3
"""Convierte padrones provinciales a CSV canónico para el validador.

Uso:
  python scripts/importar_padron.py ARBA archivo_origen.csv --out-dir padrones
  python scripts/importar_padron.py CABA archivo_origen.xlsx
  python scripts/importar_padron.py AGIP archivo_origen.rar
  python scripts/importar_padron.py EntreRios archivo_origen.txt
  python scripts/importar_padron.py SantaFe archivo_origen.csv
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules import padron_layouts  # noqa: E402
from app.modules.padrones import PADRONES_PROVINCIAS, _normalizar_row  # noqa: E402

PROVINCIAS_IMPORTABLES = {
    "ARBA": "ARBA",
    "BUENOSAIRES": "ARBA",
    "BSAS": "ARBA",
    "CABA": "CABA",
    "AGIP": "CABA",
    "CAPITAL": "CABA",
    "CAPITALFEDERAL": "CABA",
    "ENTRERIOS": "EntreRios",
    "ENTRE_RIOS": "EntreRios",
    "ATER": "EntreRios",
    "CORDOBA": "Cordoba",
    "DGRCORDOBA": "Cordoba",
    "FORMOSA": "Formosa",
    "JUJUY": "Jujuy",
    "MENDOZA": "Mendoza",
    "ATM": "Mendoza",
    "SANTAFE": "SantaFe",
    "SANTA_FE": "SantaFe",
    "APISANTAFE": "SantaFe",
    "TUCUMAN": "Tucuman",
    "TUCUMÁN": "Tucuman",
}

CANONICAL_HEADERS = [
    "cuit",
    "alicuota_retencion",
    "alicuota_percepcion",
    "vigencia_desde",
    "vigencia_hasta",
    "regimen",
]

PERFILES_CALIDAD = {
    "CABA": {
        "nombre": "AGIP/CABA · Regímenes Generales",
        "layout_esperado": {"agip_regimenes_generales", "cabeceras_alias"},
        "min_registros": 1,
        "max_caida_pct": -50,
        "permite_sin_alicuota": False,
        "requiere_vigencia": True,
    },
    "Cordoba": {
        "nombre": "Córdoba · Padrón IIBB",
        "layout_esperado": {"cordoba_iibb_delimitado_v1", "delimitado_sin_cabecera", "cabeceras_alias", "xlsx"},
        "min_registros": 1,
        "max_caida_pct": -50,
        "permite_sin_alicuota": False,
        "requiere_vigencia": False,
    },
    "Jujuy": {
        "nombre": "Jujuy · DPR alícuotas",
        "layout_esperado": {"jujuy_iibb_xlsx_alias_v1", "xlsx", "cabeceras_alias", "delimitado_sin_cabecera"},
        "min_registros": 1,
        "max_caida_pct": -50,
        "permite_sin_alicuota": False,
        "requiere_vigencia": True,
    },
    "Tucuman": {
        "nombre": "Tucumán · RG 23/02",
        "layout_esperado": {"tucuman_iibb_rg23_csv_v1", "cabeceras_alias", "delimitado_sin_cabecera"},
        "min_registros": 1,
        "max_caida_pct": -50,
        "permite_sin_alicuota": True,
        "requiere_vigencia": False,
    },
    "_default": {
        "nombre": "Perfil estándar",
        "layout_esperado": set(),
        "min_registros": 1,
        "max_caida_pct": -50,
        "permite_sin_alicuota": False,
        "requiere_vigencia": False,
    },
}


def _norm_provincia(valor: str) -> str:
    key = (
        re.sub(r"[^A-Za-zÁÉÍÓÚÜÑáéíóúüñ]", "", valor or "")
        .upper()
        .translate(str.maketrans("ÁÉÍÓÚÜÑ", "AEIOUUN"))
    )
    if key not in PROVINCIAS_IMPORTABLES:
        opciones = ", ".join(sorted(p for p, cfg in PADRONES_PROVINCIAS.items() if cfg.get("tipo") == "archivo"))
        raise SystemExit(f"Provincia no soportada: {valor}. Opciones: {opciones}")
    provincia = PROVINCIAS_IMPORTABLES[key]
    if PADRONES_PROVINCIAS[provincia].get("tipo") != "archivo":
        raise SystemExit(f"{provincia} no tiene padrón por archivo normalizado; figura como consulta manual/credenciales.")
    return provincia


def _leer_texto(path: Path) -> str:
    data = path.read_bytes()
    for enc in ("utf-8-sig", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _rows_csv(path: Path) -> list[dict]:
    texto = _leer_texto(path)
    try:
        dialect = csv.Sniffer().sniff(texto[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return list(csv.DictReader(io.StringIO(texto), dialect=dialect))


def _rows_xlsx(path: Path, sheet: str | None = None) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for idx, row in enumerate(rows[:20]):
        vals = ["" if v is None else str(v).strip() for v in row]
        joined = " ".join(vals).lower()
        if "cuit" in joined or "nro_cuit" in joined or "numero" in joined:
            header_idx = idx
            break
    if header_idx is None:
        raise SystemExit("No pude detectar fila de cabecera con CUIT en el XLSX.")
    headers = ["" if v is None else str(v).strip() for v in rows[header_idx]]
    out = []
    for row in rows[header_idx + 1:]:
        d = {headers[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(row) if i < len(headers)}
        if any(d.values()):
            out.append(d)
    return out


def _normalizar_fecha_agip(valor: str) -> str:
    digitos = _solo_digitos(valor)
    if len(digitos) == 8:
        return f"{digitos[:2]}/{digitos[2:4]}/{digitos[4:]}"
    return str(valor or "").strip()


def _normalizar_pct_agip(valor: str) -> str:
    raw = str(valor or "").replace("%", "").strip()
    if "," in raw or "." in raw:
        return _normalizar_pct(raw)
    digitos = _solo_digitos(raw)
    if len(digitos) in {3, 4}:
        return f"{int(digitos) / 100:.2f}"
    return raw


def _rows_agip_layout(path: Path) -> list[dict]:
    """Parsea el layout AGIP separado por ';' publicado como diseño de registro."""
    layout = padron_layouts.obtener_layout("agip_caba_regimenes_generales_v1")
    out = []
    for line in _leer_texto(path).splitlines():
        if not line.strip() or "cuit" in line.lower():
            continue
        row = padron_layouts.traducir_linea_delimitada(line, layout) if layout else None
        if row:
            out.append(row)
    return out


def _rows_layout_cabeceras(provincia: str, rows: list[dict], parse_meta: dict) -> list[dict]:
    out = []
    for layout in padron_layouts.layouts_para_padron(provincia):
        if layout.get("formato") != "csv_xlsx_con_cabeceras" or not layout.get("fuente_id"):
            continue
        translated = []
        for row in rows:
            parsed = padron_layouts.traducir_row_con_cabeceras(row, layout)
            if parsed:
                translated.append(parsed)
        if translated:
            parse_meta.setdefault("layout_detectado", layout["id"])
            out = translated
            break
    return out


def _rows_layout_delimitado(provincia: str, path: Path, parse_meta: dict) -> list[dict]:
    out = []
    for layout in padron_layouts.layouts_para_padron(provincia):
        if layout.get("formato") != "txt_delimitado_sin_header" or not layout.get("fuente_id"):
            continue
        translated = []
        for line in _leer_texto(path).splitlines():
            if not line.strip() or "cuit" in line.lower():
                continue
            parsed = padron_layouts.traducir_linea_delimitada(line, layout)
            if parsed:
                translated.append(parsed)
        if translated:
            parse_meta.setdefault("layout_detectado", layout["id"])
            out = translated
            break
    return out


def _rows_delimitado_sin_header(path: Path) -> list[dict]:
    """Fallback para archivos provinciales delimitados sin cabecera explícita."""
    out = []
    for line in _leer_texto(path).splitlines():
        if not line.strip():
            continue
        separador = next((sep for sep in (";", "\t", "|") if sep in line), None)
        cols = [c.strip() for c in (line.split(separador) if separador else line.split(",")) if c.strip()]
        cuit_idx = next((i for i, c in enumerate(cols) if len(_solo_digitos(c)) == 11), None)
        if cuit_idx is None:
            continue
        porcentajes = [c for c in cols[cuit_idx + 1:] if re.search(r"\d{1,2}(?:[,.]\d{1,4}|%)", c)]
        fechas = [c for c in cols[cuit_idx + 1:] if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", c)]
        out.append({
            "cuit": cols[cuit_idx],
            "alicuota_retencion": porcentajes[0] if len(porcentajes) > 0 else "",
            "alicuota_percepcion": porcentajes[1] if len(porcentajes) > 1 else "",
            "vigencia_desde": fechas[0] if len(fechas) > 0 else "",
            "vigencia_hasta": fechas[1] if len(fechas) > 1 else "",
            "regimen": "Delimitado sin cabecera",
        })
    return out


def _rows_txt_ancho_fijo(path: Path) -> list[dict]:
    """Fallback simple para TXT sin cabecera: busca CUIT y hasta dos porcentajes por línea."""
    out = []
    for line in _leer_texto(path).splitlines():
        cuit_match = re.search(r"\b\d{11}\b", line)
        if not cuit_match:
            continue
        porcentajes = re.findall(r"\d{1,2}(?:[,.]\d{1,4})", line[cuit_match.end():])
        out.append({
            "cuit": cuit_match.group(0),
            "alicuota_retencion": porcentajes[0] if len(porcentajes) > 0 else "",
            "alicuota_percepcion": porcentajes[1] if len(porcentajes) > 1 else "",
            "vigencia_desde": "",
            "vigencia_hasta": "",
            "regimen": "TXT ancho fijo",
        })
    return out


def _extraer_zip(path: Path, destino: Path) -> None:
    with zipfile.ZipFile(path) as zf:
        base = destino.resolve()
        for member in zf.infolist():
            target = (base / member.filename).resolve()
            if base != target and base not in target.parents:
                raise RuntimeError(f"ZIP inválido: ruta fuera de destino ({member.filename}).")
            zf.extract(member, base)


def _extraer_rar(path: Path, destino: Path) -> str:
    comandos = [
        ("unar", ["unar", "-quiet", "-force-overwrite", "-output-directory", str(destino), str(path)]),
        ("7z", ["7z", "x", "-y", f"-o{destino}", str(path)]),
        ("bsdtar", ["bsdtar", "-xf", str(path), "-C", str(destino)]),
        ("unrar", ["unrar", "x", "-o+", str(path), str(destino)]),
    ]
    for binario, cmd in comandos:
        if shutil.which(binario):
            try:
                subprocess.run(cmd, check=True, capture_output=True, text=True)
                return binario
            except subprocess.CalledProcessError as exc:
                detalle = (exc.stderr or exc.stdout or str(exc)).strip()
                raise RuntimeError(f"No pude extraer RAR con {binario}: {detalle}") from exc
    raise RuntimeError("Para importar RAR necesitás instalar un extractor disponible en PATH: unar, 7z, bsdtar o unrar.")


def _score_archivo_extraido(path: Path) -> tuple[int, int]:
    nombre = path.name.lower()
    ext_score = {".txt": 50, ".csv": 45, ".tsv": 40, ".psv": 40, ".xlsx": 35, ".xlsm": 35}.get(path.suffix.lower(), 0)
    name_score = sum(p in nombre for p in ("padron", "padrón", "alicuota", "regimen", "régimen", "iibb", "cuit")) * 10
    return (ext_score + name_score, path.stat().st_size)


def _archivos_extraidos(path: Path, destino: Path, parse_meta: dict) -> list[Path]:
    if path.suffix.lower() == ".zip":
        _extraer_zip(path, destino)
        parse_meta["extractor"] = "zipfile"
    elif path.suffix.lower() == ".rar":
        parse_meta["extractor"] = _extraer_rar(path, destino)
    else:
        return [path]
    soportados = {".csv", ".txt", ".tsv", ".psv", ".xlsx", ".xlsm"}
    archivos = [p for p in destino.rglob("*") if p.is_file() and p.suffix.lower() in soportados]
    if not archivos:
        raise RuntimeError(f"No encontré archivos importables dentro de {path.name}.")
    return sorted(archivos, key=_score_archivo_extraido, reverse=True)


def _leer_origen(provincia: str, path: Path, sheet: str | None, parse_meta: dict | None = None) -> list[dict]:
    parse_meta = parse_meta if parse_meta is not None else {}
    if path.suffix.lower() in {".zip", ".rar"}:
        errores = []
        with tempfile.TemporaryDirectory(prefix="padron_extract_") as tmp:
            for extraido in _archivos_extraidos(path, Path(tmp), parse_meta):
                try:
                    rows = _leer_origen(provincia, extraido, sheet, parse_meta)
                    if rows:
                        parse_meta["archivo_interno"] = extraido.name
                        return rows
                except Exception as exc:
                    errores.append(f"{extraido.name}: {exc}")
        raise RuntimeError(f"No pude normalizar ningún archivo dentro de {path.name}: {'; '.join(errores)}")

    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm"}:
        rows = _rows_xlsx(path, sheet)
        specific = _rows_layout_cabeceras(provincia, rows, parse_meta)
        if specific:
            return specific
        parse_meta.setdefault("layout_detectado", "xlsx")
        return rows
    if ext in {".csv", ".txt", ".tsv", ".psv"}:
        if provincia == "CABA":
            agip = _rows_agip_layout(path)
            if agip:
                parse_meta.setdefault("layout_detectado", "agip_regimenes_generales")
                return agip
        specific_delimited = _rows_layout_delimitado(provincia, path, parse_meta)
        if specific_delimited:
            return specific_delimited
        rows = _rows_csv(path)
        specific = _rows_layout_cabeceras(provincia, rows, parse_meta)
        if specific:
            return specific
        if rows and any((k or "").strip() for k in rows[0].keys()):
            norm = [_normalizar_row(r) for r in rows]
            if any(r.get("cuit") for r in norm):
                parse_meta.setdefault("layout_detectado", "cabeceras_alias")
                return rows
        sin_header = _rows_delimitado_sin_header(path)
        if sin_header:
            parse_meta.setdefault("layout_detectado", "delimitado_sin_cabecera")
            return sin_header
        parse_meta.setdefault("layout_detectado", "txt_ancho_fijo")
        return _rows_txt_ancho_fijo(path)
    raise SystemExit(f"Extensión no soportada: {ext}. Usá CSV, TXT, XLSX, ZIP o RAR.")


def _solo_digitos(valor: str) -> str:
    return "".join(ch for ch in str(valor or "") if ch.isdigit())


def _normalizar_pct(valor: str) -> str:
    valor = str(valor or "").replace("%", "").replace(",", ".").strip()
    return valor


def _pct_valido(valor: str) -> bool:
    if not valor:
        return True
    try:
        numero = float(str(valor).replace(",", ".").replace("%", "").strip())
    except ValueError:
        return False
    return 0 <= numero <= 100


def _fecha_valida(valor: str) -> bool:
    if not valor:
        return True
    from app.modules.padron_manifest import _parse_fecha
    return _parse_fecha(str(valor)) is not None


def normalizar_rows_con_calidad(
    rows: Iterable[dict],
    destino: Path | None = None,
    provincia: str | None = None,
    layout_detectado: str = "",
) -> tuple[list[dict], dict]:
    perfil = PERFILES_CALIDAD.get(provincia or "", PERFILES_CALIDAD["_default"])
    out = []
    vistos = set()
    raw_registros = 0
    descartados_cuit_invalido = 0
    duplicados = 0
    alicuotas_invalidas = 0
    fechas_invalidas = 0
    for row in rows:
        raw_registros += 1
        n = _normalizar_row(row)
        cuit = _solo_digitos(n.get("cuit", ""))
        if len(cuit) != 11:
            descartados_cuit_invalido += 1
            continue
        if cuit in vistos:
            duplicados += 1
            continue
        vistos.add(cuit)
        ret = _normalizar_pct(n.get("alicuota_retencion", ""))
        perc = _normalizar_pct(n.get("alicuota_percepcion", ""))
        desde = n.get("vigencia_desde", "")
        hasta = n.get("vigencia_hasta", "")
        if not _pct_valido(ret) or not _pct_valido(perc):
            alicuotas_invalidas += 1
        if not _fecha_valida(desde) or not _fecha_valida(hasta):
            fechas_invalidas += 1
        out.append({
            "cuit": cuit,
            "alicuota_retencion": ret,
            "alicuota_percepcion": perc,
            "vigencia_desde": desde,
            "vigencia_hasta": hasta,
            "regimen": n.get("regimen", ""),
        })
    sin_alicuota = sum(1 for r in out if not r["alicuota_retencion"] and not r["alicuota_percepcion"])
    registros_previos = _contar_csv(destino) if destino else None
    variacion_pct = None
    if registros_previos:
        variacion_pct = round(((len(out) - registros_previos) / registros_previos) * 100, 2)
    errores = []
    advertencias = []
    if len(out) < perfil["min_registros"]:
        errores.append("No se detectaron CUITs válidos para importar.")
    if descartados_cuit_invalido:
        advertencias.append(f"{descartados_cuit_invalido} filas descartadas por CUIT inválido.")
    if duplicados:
        advertencias.append(f"{duplicados} CUITs duplicados descartados.")
    if sin_alicuota and not perfil["permite_sin_alicuota"] and sin_alicuota == len(out):
        advertencias.append("Todos los registros normalizados quedaron sin alícuotas.")
    elif sin_alicuota and not perfil["permite_sin_alicuota"]:
        advertencias.append(f"{sin_alicuota} registros quedaron sin alícuotas.")
    if alicuotas_invalidas:
        advertencias.append(f"{alicuotas_invalidas} registros tienen alícuotas no numéricas o fuera de rango.")
    if fechas_invalidas:
        advertencias.append(f"{fechas_invalidas} registros tienen fechas no parseables.")
    if perfil["requiere_vigencia"] and out and all(not r["vigencia_desde"] and not r["vigencia_hasta"] for r in out):
        advertencias.append(f"El perfil {perfil['nombre']} espera vigencia informada.")
    if perfil["layout_esperado"] and layout_detectado and layout_detectado not in perfil["layout_esperado"]:
        advertencias.append(f"Layout detectado '{layout_detectado}' fuera del perfil esperado para {perfil['nombre']}.")
    if variacion_pct is not None and variacion_pct <= perfil["max_caida_pct"]:
        advertencias.append(f"El nuevo padrón tiene una caída de {abs(variacion_pct)}% contra el archivo vigente.")
    estado = "bloqueado" if errores else ("observado" if advertencias else "aprobado")
    calidad = {
        "estado": estado,
        "raw_registros": raw_registros,
        "registros_validos": len(out),
        "descartados_cuit_invalido": descartados_cuit_invalido,
        "duplicados": duplicados,
        "sin_alicuota": sin_alicuota,
        "alicuotas_invalidas": alicuotas_invalidas,
        "fechas_invalidas": fechas_invalidas,
        "registros_previos": registros_previos,
        "variacion_pct": variacion_pct,
        "perfil": perfil["nombre"],
        "layout_detectado": layout_detectado,
        "errores": errores,
        "advertencias": advertencias,
    }
    return out, calidad


def normalizar_rows(rows: Iterable[dict]) -> list[dict]:
    out, _calidad = normalizar_rows_con_calidad(rows)
    return out


def _contar_csv(path: Path | None) -> int | None:
    if not path or not path.exists():
        return None
    with path.open(encoding="utf-8", newline="") as f:
        return sum(1 for _ in csv.DictReader(f))


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _evidencia_origen(origen: Path, parse_meta: dict) -> dict:
    return {
        "archivo_original": origen.name,
        "tipo_archivo": origen.suffix.lower(),
        "tamano_bytes": origen.stat().st_size,
        "sha256": _sha256(origen),
        "extractor": parse_meta.get("extractor", ""),
        "archivo_interno": parse_meta.get("archivo_interno", ""),
        "layout_detectado": parse_meta.get("layout_detectado", ""),
    }


def escribir_csv(rows: list[dict], destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)
    with destino.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CANONICAL_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def importar_padron(
    provincia_arg: str,
    origen: Path,
    out_dir: Path,
    sheet: str | None = None,
    dry_run: bool = False,
    periodo: str | None = None,
    vigencia_hasta: str | None = None,
    backup: bool = True,
    aceptar_observado: bool = True,
) -> dict:
    provincia = _norm_provincia(provincia_arg)
    cfg = PADRONES_PROVINCIAS[provincia]
    origen = origen.expanduser().resolve()
    if not origen.exists():
        raise FileNotFoundError(f"No existe el archivo origen: {origen}")

    destino = out_dir / cfg["archivo"]
    parse_meta: dict = {}
    rows_raw = _leer_origen(provincia, origen, sheet, parse_meta)
    rows, calidad = normalizar_rows_con_calidad(rows_raw, destino, provincia, parse_meta.get("layout_detectado", ""))
    evidencia = _evidencia_origen(origen, parse_meta)
    if calidad["estado"] == "bloqueado":
        raise RuntimeError("; ".join(calidad["errores"]))
    if calidad["estado"] == "observado" and not dry_run and not aceptar_observado:
        raise RuntimeError("La importación tiene advertencias. Previsualizá y confirmá explícitamente para sobrescribir el padrón.")
    backup_path = None
    if not dry_run:
        from app.modules.padron_manifest import backup_si_existe, registrar_carga
        if backup:
            backup_path = backup_si_existe(destino)
        escribir_csv(rows, destino)
        registrar_carga(out_dir, provincia, cfg["archivo"], len(rows), str(origen), periodo, vigencia_hasta, backup_path, evidencia, calidad)
    return {
        "provincia": provincia,
        "nombre": cfg["nombre"],
        "origen": str(origen),
        "destino": str(destino),
        "registros": len(rows),
        "muestra": rows[:3],
        "dry_run": dry_run,
        "periodo": periodo or "",
        "vigencia_hasta": vigencia_hasta or "",
        "backup": backup_path,
        "calidad": calidad,
        "evidencia": evidencia,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa padrones provinciales al formato canónico del validador.")
    parser.add_argument("provincia", help="Provincia con padrón por archivo: ARBA, CABA, EntreRios, Cordoba, Formosa, Jujuy, Mendoza, SantaFe, Tucuman")
    parser.add_argument("origen", type=Path, help="Archivo origen CSV/TXT/XLSX/ZIP/RAR")
    parser.add_argument("--out-dir", type=Path, default=ROOT / "padrones", help="Carpeta destino")
    parser.add_argument("--sheet", help="Hoja XLSX a leer")
    parser.add_argument("--periodo", help="Período del padrón, ej. 2026-06")
    parser.add_argument("--vigencia-hasta", help="Fecha de vigencia hasta, ej. 2026-06-30")
    parser.add_argument("--sin-backup", action="store_true", help="No respalda el padrón anterior")
    parser.add_argument("--dry-run", action="store_true", help="No escribe archivo; solo informa")
    args = parser.parse_args()

    try:
        resultado = importar_padron(
            args.provincia,
            args.origen,
            args.out_dir,
            args.sheet,
            args.dry_run,
            args.periodo,
            args.vigencia_hasta,
            not args.sin_backup,
        )
    except Exception as e:
        raise SystemExit(str(e)) from e

    print(f"Provincia: {resultado['provincia']} ({resultado['nombre']})")
    print(f"Origen: {resultado['origen']}")
    print(f"Registros normalizados: {resultado['registros']}")
    print(f"Destino: {resultado['destino']}")
    if resultado["muestra"]:
        print("Muestra:")
        for row in resultado["muestra"]:
            print("  " + ", ".join(f"{k}={row[k]}" for k in CANONICAL_HEADERS))
    if not args.dry_run:
        print("OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
